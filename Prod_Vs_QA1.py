#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 15 20:32:51 2024

@author: charanmakkina
"""

import pandas as pd
import numpy as np
import glob
import os
import re
import sys

#prod_path=sys.argv[1]+'\\'
#qa_path=sys.argv[2]+'\\'
#roundoff=sys.argv[3]
#roundoff=int(roundoff)
roundoff=8
path_templates=r'Templates\\'
prod_path=r'/Users/charanmakkina/Python_Spyder/Comparisons/prod//'
qa_path=r'/Users/charanmakkina/Python_Spyder/Comparisons/prod//'
prod_file='*CalculationResults*.xlsx'
qa_file='*CalculationResults*.xlsx'
archive_folder=os.path.join(prod_path, 'Comparison')
if not os.path.exists(archive_folder):
    os.mkdir(archive_folder)

file_mapping_prod=prod_path+prod_file
file_mapping_qa=qa_path+qa_file 
file_paths_prod=glob.glob(file_mapping_prod)
file_paths_qa=glob.glob(file_mapping_qa)

file_name_prod=[]
file_name_qa=[]
for file_path_prod in file_paths_prod:
    file_name_prod.append(os.path.basename(file_path_prod))

for file_path_qa in file_paths_qa:
    file_name_qa.append(os.path.basename(file_path_qa))

prod_id=[]

for file in file_name_prod:
    prod_id.append('_'.join(file.split('_')[:8]))

qa_id=[]

for file in file_name_qa:
    qa_id.append('_'.join(file.split('_')[:8]))

prod_df=pd.DataFrame({'Id':prod_id,'prod_files': file_name_prod}) 
prod_df.drop_duplicates(subset='Id',keep='first', inplace=True)

qa_df=pd.DataFrame({'Id':qa_id, 'qa_files':file_name_qa}) 
qa_df.drop_duplicates(subset='Id', keep='first', inplace=True)

df=prod_df.merge(qa_df,how='inner',on='Id') 
df.set_index('Id',inplace=True)

comp=[]
comp_all=[]
prod=[]
qa=[]
dp_ids=[]
prod_only=[]
qa_only=[]
trades_priced=[]
price_match=[]
status=[]
num_match=[]

for prod_file,qa_file,dp_id in zip(df['prod_files'], df['qa_files'], df.index):
    df_prod=pd.read_excel(prod_path+prod_file) 
    df_qa=pd.read_excel(qa_path+qa_file)
    
    #df_prod=df_prod.apply(lambda column: column.astype(str) if column.isnull().all() else column) 
    #df_qa=df_qa.apply(lambda column: column.astype(str) if column.isnull().all() else column) 
    merge_columns={'Trade':['Trade ID','Valuation Status']}
    key_id='Trade'
    prod_columns=[col for col in df_prod.columns if col in df_qa.columns]
    df_prod=df_prod[prod_columns].copy()
    
    df_merge_left=df_prod.merge(df_qa,on=merge_columns[key_id], how='left', indicator=True, suffixes=('','_qa'))
    
    df_only_prod=df_merge_left[df_merge_left['_merge']=='left_only'] 
    df_only_prod=df_only_prod[df_prod.columns]
    
    
    df_merge_right=df_prod.merge(df_qa, on=merge_columns[key_id], how='right', indicator=True, suffixes=('_prod',''))
    df_only_qa=df_merge_right[df_merge_right['_merge']=='right_only']
    df_only_qa=df_only_qa[df_qa.columns]
    
    
    df_both=df_merge_left[df_merge_left['_merge']=='both']
    
    df_prod.set_index(merge_columns[key_id], inplace=True) 
    df_qa.set_index(merge_columns[key_id], inplace=True)
    
    df_only_prod.set_index(merge_columns[key_id], inplace=True)
    df_only_qa.set_index(merge_columns[key_id], inplace=True)
    
    df_comp_num=df_prod.select_dtypes(include=['number']).sub(df_qa[df_prod.select_dtypes(include=['number']).columns])
    df_comp_cat=df_prod.select_dtypes(include=['object','category']).replace(np.nan,'').eq(df_qa[df_prod.select_dtypes(include=['object','category']).columns].replace(np.nan,''))
    df_comp=pd.concat([df_comp_num,df_comp_cat],axis=1)
    df_comp.reset_index(inplace=True)
    #df_comp=df_comp[df_comp[merge_columns[key_id]].isin(df_both[merge_columns[key_id]])]
    df_prod.reset_index(inplace=True)
    df_comp=df_comp.merge(df_both,how='inner',on=merge_columns[key_id],suffixes=('','_both'))
    df_comp=df_comp[df_prod.columns]
    
    df_comp.set_index(merge_columns[key_id], inplace=True) 
    
    df_comp_num.reset_index(inplace=True)
    df_comp_num_copy=df_comp_num.copy()
    df_comp_num=df_comp_num.merge(df_both,how='inner',on=merge_columns[key_id],suffixes=('','_both'))
    df_comp_num=df_comp_num[df_comp_num_copy.columns]
    #df_comp_num=df_comp_num[df_comp_num[merge_columns[key_id]].isin(df_both[merge_columns[key_id]])]
    df_comp_num.set_index(merge_columns[key_id], inplace=True)
    df_prod.set_index(merge_columns[key_id],inplace=True)
    #total-pd.DataFrame (index-['Total'],columns-df_comp.columns)
    
    status_cat=pd.DataFrame((df_comp[df_comp_cat.columns]==True).all()).T
    
    status_overall=(status_cat.T==True).all().iloc[0] 
    status_cat_T=(status_cat.copy().T)
    
    status_cat_T=status_cat_T[status_cat_T[0]!=True].index 
    status_list=','.join(status_cat_T)
    
    columns_with_all_na=df_comp.columns[df_comp.isna().all()].to_list()
    columns_all_na=df_comp.columns[df_comp.isna().all()]
    
    status_num=pd.DataFrame((df_comp_num.eq(0).sum(skipna=True))).T
    status_num_dec=pd.DataFrame((round(df_comp_num, int(roundoff)).eq(0).sum(skipna=True))).T
    status_num_count=pd.DataFrame((df_comp_num).count()).T
    status_num[columns_with_all_na]=np.nan 
    status_num_dec[columns_with_all_na]=np.nan
    status_num_count[columns_with_all_na]=np.nan
    status_num_count=status_num_count.T
    status_num_count=status_num_count[~status_num_count[0].isna()]
    total=pd.concat([status_cat, status_num], axis=1)
    total['index']='Total/ZeroDiffTrades'
    total.set_index('index', inplace=True)
    
    #total=total.replace(int(0), np.nan)
    
    numerical_match_columns=df_prod.select_dtypes(include=['number']).columns
    numerical_match_columns=[col for col in numerical_match_columns if col not in columns_all_na] 
    numerical_match=pd.DataFrame(total.loc['Total/ZeroDiffTrades',numerical_match_columns]).T
    numerical_match_percent=(numerical_match.T.rename(columns={'Total/ZeroDiffTrades':0}))/status_num_count
    numerical_match_percent=numerical_match_percent.T
    total_dec=pd.concat([status_cat, status_num_dec], axis=1)
    total_dec['index']='Total/ZeroDiffTradesdec_roundoff'
    total_dec.set_index('index', inplace=True)
    #total dec-total dec.replace(int(0) e(int(0), np. .nan)
    numerical_match_dec=pd.DataFrame(total_dec.loc[ 'Total/ZeroDiffTradesdec_roundoff', numerical_match_columns]).T
    numerical_match_dec_percent=(numerical_match_dec.T.rename(columns={"Total/ZeroDiffTradesdec_roundoff":0}))/status_num_count
    numerical_match_dec_percent=numerical_match_dec_percent.T

#-status

    df_prod_c=df_prod.copy()
    
    df_qa_c=df_qa.copy() 
    df_comp_c=df_comp.copy()
    
    df_prod_c.reset_index(inplace=True)
    
    df_qa_c.reset_index(inplace=True) 
    df_comp_c.reset_index(inplace=True)
    
    df_prod_c=df_prod_c[df_prod_c[merge_columns[key_id]].isin(df_comp_c[merge_columns[key_id]])]
    df_qa_c=df_qa_c[df_qa_c[merge_columns[key_id]].isin(df_comp_c[merge_columns[key_id]])]
    
    df_prod_c.set_index(merge_columns[key_id], inplace=True) 
    df_qa_c.set_index(merge_columns[key_id], inplace=True)
    
    df_comp_c.set_index(merge_columns[key_id], inplace=True)
    
    df_prod_c.columns=pd.MultiIndex.from_product([df_prod_c.columns,['prod']],names=['outer', 'inner'])
    df_qa_c.columns=pd.MultiIndex.from_product([df_qa_c.columns,['qa']],names=['outer', 'inner'])
    df_comp_c.columns=pd.MultiIndex.from_product([df_comp_c.columns,['diff']],names=['outer', 'inner'])
    df_prod_qa_comp=pd.concat([df_prod_c,df_qa_c,df_comp_c],axis=1)
    
    
    sort_order=df_prod.columns.tolist()
    
    df_prod_qa_comp=df_prod_qa_comp.reindex(columns=sort_order, level=0)    
    df_comp=df_comp[df_prod.columns]
    total=total[df_comp.columns]
    total_dec=total_dec[df_comp.columns]
    df_comp=pd.concat([df_comp, total, total_dec],axis=0)
    
    dp_ids.append(dp_id) 
    comp.append(df_comp)
    comp_all.append(df_prod_qa_comp) 
    prod.append(df_prod)
    qa.append(df_qa)
    
    prod_only.append(df_only_prod) 
    qa_only.append(df_only_qa)
    
    
    status.append([status_overall, status_list])
    
    num_match.append([numerical_match,numerical_match_dec, numerical_match_percent, numerical_match_dec_percent])
for ids, files, comb,p,q,po, qo in zip(dp_ids,comp,comp_all,prod,qa,prod_only,qa_only):
    with pd.ExcelWriter(archive_folder+'//comp'+ids+'.xlsx',engine='xlsxwriter') as writer:
        files.to_excel(writer, sheet_name='comparison', index=True) 
        comb.to_excel(writer, sheet_name='prod_qa_comp', index=True)
        p.to_excel(writer, sheet_name='Prod', index=True)
        q.to_excel(writer, sheet_name='QA', index=True) 
        po.to_excel(writer, sheet_name='Prod only', index=True)
        
        qo.to_excel(writer, sheet_name='QA_only', index=True)
        
df_kpi=[]

df_num=pd.DataFrame()

df_num_dec=pd.DataFrame()
        
for ids, files,p,q,po, qo, s, nm in zip(dp_ids,comp,prod,qa,prod_only,qa_only, status, num_match):

    date=ids.split('_')[0]
    
    SNAP='_'.join(ids.split('_')[1:3]) 
    Curve='_'.join(ids.split('_')[3:6])
    Asset=ids.split('_')[6]
    
    df_kpi.append({'Dote': date, 'SNAP': SNAP, 'Curve': Curve, 'Asset':Asset, 'Prod':len(p), 
                   'QA':len(q), 'Both': len(files)-2, 'Prod Only':len(po), 'QA only': len(qo), 
                   'Status Cat':s[0], 'Not Matched Cat Fields':s[1]})
    
    #df_kpi.append([nm[0],nm[1]]) 
    if(df_num.size==0):
        df_num=nm[0]
        df_num_dec=nm[1]
    
    else:
        df_num=pd.concat([df_num,nm[0]])
        df_num_dec=pd.concat([df_num_dec,nm[1]])
        
df_num.reset_index(inplace=True, drop=True) 
df_num_dec.reset_index(inplace=True, drop=True)

kpi=pd.DataFrame(df_kpi)

kpi_num=pd.concat([kpi,df_num],axis=1)

kpi_num_dec=pd.concat([kpi,df_num_dec],axis=1) 
kpi_num.to_excel(archive_folder+'//KPI_num.xlsx', index=False)

kpi_num_dec.to_excel(archive_folder+'//KPI_num_dec.xlsx',index=False)

df_pkpi=[]

df_num=pd.DataFrame()

df_num_dec=pd.DataFrame()

for ids, files,p,q,po, qo, s, nm in zip(dp_ids,comp,prod,qa,prod_only,qa_only, status, num_match):

    date=ids.split('_')[0]
    
    SNAP='_'.join(ids.split('_')[1:3]) 
    Curve='_'.join(ids.split('_')[3:6])
    Asset=ids.split('_')[6]

    both=len(files)-2
    
    df_pkpi.append({'Date':date, 'SNAP': SNAP, 'Curve': Curve, 'Asset':Asset, 'Prod':len(p), 
                    'QA':len(q), 'Both': len(files)-2, 'Prod Only':len(po), 'QA_only':len(qo),
                    'Status Cat':s[0], 'Not Matched Cat Fields':s[1]})
    
    #pkpi=pd.DataFrame(df_pkp1)
    
    #pkpi.to_excel(archive_folder+'\\KPI_Percent.xlsx', Index-False)
    
    if(df_num.size==0):
        df_num=nm[2]
        df_num_dec=nm[3] 
    else:
    
        df_num=pd.concat([df_num,nm[2]])
        
        df_num_dec=pd.concat([df_num_dec,nm[3]])

df_num.reset_index(inplace=True, drop=True) 
df_num_dec.reset_index(inplace=True, drop=True)

kpi=pd.DataFrame(df_kpi) 
kpi_num=pd.concat([kpi,df_num], axis=1)

kpi_num_dec=pd.concat([kpi, df_num_dec], axis=1)

kpi_num.to_excel(archive_folder+'//KPI_percent_num.xlsx',index=False)
kpi_num_dec.to_excel(archive_folder+'//KPI_percent_num_dec.xlsx', index=False)

import openpyxl

def copy_formatting(source_file, source_sheet, target_file, target_sheet):

    #Load source and target workbooks

    source_wb=openpyxl.load_workbook(source_file, read_only=True) 
    target_wb=openpyxl.load_workbook(target_file)
    
    #Select source and target sheets
    source_ws=source_wb[source_sheet] 
    target_ws=target_wb[target_sheet]
    
    #Copy formatting from row 1 in source sheet to row 1 in the target sheet
    
    source_row_1=source_ws[1] 
    target_row_1=target_ws[1]
    
    for source_cell, target_cell in zip(source_row_1, target_row_1): 
        #Copy styles (formatting) from the source cell to the target cell 
        target_cell.font=source_cell.font
        target_cell.border=source_cell.border
        target_cell.fill=source_cell.fill
        target_cell.number_format=source_cell.number_format
        target_cell.alignment=source_cell.alignment
        
    #Copy formatting from row 2 in the source sheet to all other rows in the target sheet
    
    source_row_2=source_ws[2]
    
    for target_row in target_ws.iter_rows(min_row=2, max_row=target_ws.max_row):
        for source_cell, target_cell in zip(source_row_2, target_row):
    
            #Copy styles (formatting) from the source cell to the target cell 
            target_cell.font=source_cell.font 
            target_cell.border=source_cell.border
            target_cell.fill=source_cell.fill 
            target_cell.number_format=source_cell.number_format
            target_cell.alignment=source_cell.alignment
            
    #Save the changes to the target workbook
    
    target_wb.save(target_file)

#Example usage


#copy_formatting(path_templates+ 'KPI - Template2.xlsx', 'Sheet1', archive_folder+'//KPI_num.xlsx', 'Sheet1')

#copy_formatting(path_templates+ 'KPI - Template2.xlsx', 'Sheet1', archive_folder+'//KPI_num_dec.xlsx', 'Sheet1')

#copy_formatting(path_templates+ 'KPI_Percent - Template2.xlsx', 'Sheet1', archive_folder+'//KPI_percent_num.xlsx', 'Sheet1')

#copy_formatting(path_templates+ 'KPI_Percent - Template2.xlsx', 'Sheet1', archive_folder+'//KPI_percent_num_dec.xlsx', 'Sheet1')
